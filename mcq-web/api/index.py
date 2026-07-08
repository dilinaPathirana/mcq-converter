"""
Flask app: drag-and-drop web front end for the Sinhala legacy-font MCQ
PDF -> Excel converter, deployable as a Vercel Function.

Two modes, mirroring the desktop tool's two GUI tabs:
  POST /api/convert           - one PDF -> plain Q/Option Excel
  POST /api/convert-template  - Sinhala (required) + English/Tamil (optional)
                                 PDFs -> HTML-wrapped multi-language template

Vercel's Python runtime auto-detects this Flask app because it defines a
top-level `app` variable in an entrypoint file (api/index.py).

Everything is inlined into this single file on purpose: Vercel's Python
runtime loads api/index.py directly (not as part of a package), so a
sibling "from mcq_core import ..." style import fails at runtime with
ModuleNotFoundError even though the file sits right next to it. Keeping
one self-contained file sidesteps that entirely.
"""

import io
from flask import Flask, request, send_file, Response

import re

SINHALA_FONTS = {"FMAbhayax", "FMSamanthax", "FMGanganeex", "FMAbhaya", "FMBindumathi"}

SUP_START = "\x01"
SUP_END = "\x02"

DEGREE_FIX_RE = re.compile(r"(\d+)0([CF])(?![a-zA-Z0-9])")


def fix_degree_symbol(text):
    return DEGREE_FIX_RE.sub(r"\1" + chr(0xB0) + r"\2", text)


def split_superscript_runs(text):
    if SUP_START not in text:
        return [(False, text)] if text else []
    parts = []
    buf = ""
    in_sup = False
    for ch in text:
        if ch == SUP_START:
            if buf:
                parts.append((in_sup, buf))
            buf = ""
            in_sup = True
        elif ch == SUP_END:
            if buf:
                parts.append((in_sup, buf))
            buf = ""
            in_sup = False
        else:
            buf += ch
    if buf:
        parts.append((in_sup, buf))
    return parts


def strip_superscript_markers(text):
    return text.replace(SUP_START, "").replace(SUP_END, "")


def extract_from_pdf_bytes(pdf_bytes):
    import fitz
    from pandukabhaya import Converter

    conv = Converter("fm_abhaya")
    pua_re = re.compile("[" + chr(0xE000) + "-" + chr(0xF8FF) + "]")

    def convert_text(text, font):
        return conv.convert(text) if font in SINHALA_FONTS else text

    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    lines = []
    for page in doc:
        d = page.get_text("dict")
        blocks = sorted(d["blocks"], key=lambda b: (round(b["bbox"][1], 1), b["bbox"][0]))
        for b in blocks:
            for l in b.get("lines", []):
                spans = l.get("spans", [])
                n = len(spans)
                parts = []
                i = 0
                while i < n:
                    s = spans[i]
                    # See convert_mcq.py (desktop tool) for why this excludes
                    # Sinhala-font spans: the superscript flag is occasionally
                    # set on normal, full-size legacy-font text too (a table
                    # layout artifact), so only trust it for plain Latin spans.
                    is_sup = bool(s.get("flags", 0) & 1) and s["font"] not in SINHALA_FONTS
                    raw = s["text"]
                    text = convert_text(raw, s["font"])
                    if is_sup and raw.strip() == "0":
                        nxt = spans[i + 1]["text"] if i + 1 < n else ""
                        if nxt.lstrip().startswith(("C", "F")):
                            parts.append(chr(0xB0))
                            i += 1
                            continue
                    if is_sup and text.strip():
                        parts.append(SUP_START + text + SUP_END)
                    else:
                        parts.append(text)
                    i += 1
                line_text = pua_re.sub("", "".join(parts))
                if line_text.strip():
                    lines.append(line_text)
            lines.append("")
    doc.close()
    return fix_degree_symbol("\n".join(lines))


NEW_LINE_RE = re.compile(r"^(\d{1,2}\.|\(\d\)|\([A-E]\)|[A-E]\s*$|[A-E]\s*-)")


def reflow(raw_text):
    lines = raw_text.split("\n")
    logical, buf = [], ""
    for l in lines:
        s = l.strip()
        if s == "":
            if buf:
                logical.append(buf)
                buf = ""
            continue
        if NEW_LINE_RE.match(s):
            if buf:
                logical.append(buf)
            buf = s
        else:
            buf = f"{buf} {s}" if buf else s
    if buf:
        logical.append(buf)
    fixed = []
    for l in logical:
        if len(re.findall(r"\(\d\)", l)) >= 2:
            fixed.extend(p.strip() for p in re.split(r"(?=\(\d\)\s)", l) if p.strip())
        else:
            fixed.append(l)
    return fixed


Q_START_RE = re.compile(r"^(\d{1,2})\.\s*(.*)$")
OPT_NUM_RE = re.compile(r"^\((\d)\)\s*(.*)$")
OPT_LET_RE = re.compile(r"^\(([A-E])\)\s*(.*)$")


def new_q(num):
    return {"num": num, "stem_lines": [], "num_opts": {}, "let_opts": {}, "extra": [], "note": ""}


def find_combo_note(raw_text):
    m = re.search(r"\d{1,2}\s*[-" + chr(0x2013) + r"]\s*\d{1,2}[^\n]{0,60}?(?:" + chr(0x0db4) + chr(0x0daf) + chr(0x0db1) + chr(0x0db8) + "|based)", raw_text)
    if not m:
        return raw_text, None
    nums = re.match(r"(\d{1,2})\s*[-" + chr(0x2013) + r"]\s*(\d{1,2})", raw_text[m.start():])
    lo, hi = int(nums.group(1)), int(nums.group(2))
    tail = raw_text[m.start():m.start() + 400]
    m2 = re.search(r"\n\s*\d{1,2}\.\s", tail[10:])
    if m2:
        tail = tail[:10 + m2.start()]
    note = re.sub(r"\s+", " ", tail).strip()
    cleaned = raw_text[:m.start()] + raw_text[m.start() + len(tail):]
    return cleaned, (lo, hi, note)


def parse_questions(lines, max_q=60, combo_note=None):
    start = 0
    for i, l in enumerate(lines):
        m = Q_START_RE.match(l)
        if m and int(m.group(1)) == 1:
            start = i
            break
    lines = lines[start:]

    questions = {}
    order = []
    cur = None

    i = 0
    while i < len(lines):
        l = lines[i]
        m = Q_START_RE.match(l)
        if m and 1 <= int(m.group(1)) <= max_q:
            num = int(m.group(1))
            cur = new_q(num)
            questions[num] = cur
            order.append(num)
            rest = m.group(2).strip()
            if rest:
                cur["stem_lines"].append(rest)
            i += 1
            continue
        mo = OPT_NUM_RE.match(l)
        if mo and cur:
            key = mo.group(1)
            if key in cur["num_opts"] and cur["num_opts"][key] != mo.group(2).strip():
                prev_idx = order.index(cur["num"]) - 1
                prev_num = order[prev_idx] if prev_idx >= 0 else None
                if prev_num is not None and not questions[prev_num]["num_opts"]:
                    questions[prev_num]["num_opts"] = dict(cur["num_opts"])
                    questions[prev_num]["note"] += " [auto-recovered option table from layout - please verify]"
                cur["num_opts"] = {}
            cur["num_opts"][key] = mo.group(2).strip()
            i += 1
            continue
        mol = OPT_LET_RE.match(l)
        if mol and cur:
            cur["let_opts"][mol.group(1)] = mol.group(2).strip()
            i += 1
            continue
        if cur:
            cur["extra"].append(l)
        i += 1

    for num, q in questions.items():
        extra_q = list(q["extra"])

        letter_lines = [e for e in extra_q if re.match(r"^[A-E]\s*-\s*", e)]
        if letter_lines:
            rest = [e for e in extra_q if e not in letter_lines]
            q["premises"] = letter_lines
            if rest:
                q["stem_lines"] = [rest[-1]]
                extra_q = rest[:-1]
            else:
                extra_q = []
            q["note"] += " [auto-detected premise list - please verify]"

        if extra_q and q["num_opts"]:
            keys = sorted(q["num_opts"].keys(), key=int)
            for k in keys:
                val = q["num_opts"][k]
                guard = 0
                while val and not strip_superscript_markers(val).rstrip().endswith((".", "?", "!")) and extra_q and guard < 5:
                    val = val.rstrip() + " " + extra_q.pop(0)
                    guard += 1
                q["num_opts"][k] = val
            if extra_q:
                q["stem_lines"] += extra_q
            extra_q = []
        q["extra"] = extra_q

        if q["let_opts"] and not q["num_opts"]:
            q["premises"] = [f"({k}) {v}" for k, v in sorted(q["let_opts"].items())]
            if combo_note and combo_note[0] <= num <= combo_note[1]:
                q["note"] += " [combo-answer question - see Notes for shared answer key]"
                q["notes_full"] = combo_note[2]
            else:
                q["note"] += " [NEEDS REVIEW: lettered options with no numbered answer choices found]"

        if not q["num_opts"] and not q["let_opts"]:
            q["note"] += " [NEEDS REVIEW: no options detected]"

    return questions, order


def extract_questions_from_pdf_bytes(pdf_bytes, max_q=60):
    """Full pipeline: PDF bytes -> (questions dict, order list)."""
    raw = extract_from_pdf_bytes(pdf_bytes)
    raw, combo_note = find_combo_note(raw)
    lines = reflow(raw)
    return parse_questions(lines, max_q=max_q, combo_note=combo_note)


def _rich_or_plain(text):
    if SUP_START not in text:
        return text
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont

    sup_font = InlineFont(vertAlign="superscript")
    runs = []
    for is_sup, chunk in split_superscript_runs(text):
        if is_sup:
            runs.append(TextBlock(sup_font, chunk))
        else:
            runs.append(chunk)
    return CellRichText(runs)


def build_simple_excel(questions, order):
    """Build the plain Q No / Question / Option 1-5 / Notes workbook.
    Returns raw xlsx bytes."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "MCQ"

    SINHALA_FONT = "Iskoola Pota"
    HEADER_FILL = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
    HEADER_FONT = Font(name=SINHALA_FONT, size=11, bold=True, color="FFFFFF")
    BODY_FONT = Font(name=SINHALA_FONT, size=11)
    THIN = Side(style="thin", color="D9D9D9")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    headers = ["Q No", "Question", "Option 1", "Option 2", "Option 3", "Option 4", "Option 5", "Notes"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for i, w in enumerate([6, 55, 22, 22, 22, 22, 22, 45], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for num in order:
        q = questions[num]
        stem = " ".join(q["stem_lines"]).strip()
        question_text = stem
        if q.get("premises"):
            question_text += "\n" + "\n".join(q["premises"])
        opts = q["num_opts"]
        note = q.get("notes_full", "") or q.get("note", "").strip()
        row_values = [num, question_text, opts.get("1", ""), opts.get("2", ""), opts.get("3", ""),
                      opts.get("4", ""), opts.get("5", ""), note]
        row_values = [_rich_or_plain(v) if isinstance(v, str) else v for v in row_values]
        ws.append(row_values)
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(
                horizontal="center" if c == 1 else "left", vertical="top", wrap_text=True
            )
        ws.row_dimensions[r].height = 60

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Multi-language HTML template mode (mirrors convert_mcq_template.py)
# ---------------------------------------------------------------------------

TEMPLATE_HEADERS = [
    "AL/OL", "Subject",
    "Sinhala Question", "Sinhala Option 1", "Sinhala Option 2", "Sinhala Option 3",
    "Sinhala Option 4", "Sinhala Option 5", "Sinhala Explanation",
    "English Question", "English Option 1", "English Option 2", "English Option 3",
    "English Option 4", "English Option 5", "English Explanation",
    "Tamil Question", "Tamil Option 1", "Tamil Option 2", "Tamil Option 3",
    "Tamil Option 4", "Tamil Option 5", "Tamil Explanation",
    "Type", "Answer", "Difficulty Level",
    "Sinhala Module", "Sinhala sub module", "Sinhala sub sub module", "Tags",
    "Notes",
]


def sup_to_html(text):
    if SUP_START not in text:
        return text
    out = []
    for is_sup, chunk in split_superscript_runs(text):
        out.append(f"<sup>{chunk}</sup>" if is_sup else chunk)
    return "".join(out)


def html_wrap(text):
    text = sup_to_html((text or "").strip())
    return f"<p>{text}</p>" if text else ""


def question_html(q):
    parts = [html_wrap(s) for s in q.get("stem_lines", []) if s.strip()]
    premises = q.get("premises")
    if premises:
        cleaned = []
        for p in premises:
            m = re.match(r"^\(?([A-E])\)?\s*-?\s*(.*)$", p)
            if m:
                cleaned.append(f"{m.group(1)} - {m.group(2)}")
            else:
                cleaned.append(p)
        parts.append(html_wrap(", ".join(cleaned)))
    return "".join(parts)


def option_html(q, key):
    return html_wrap(q.get("num_opts", {}).get(key, ""))


def build_template_rows(si_q, si_order, en_q, ta_q, level, subject):
    rows = []
    for num in si_order:
        sq = si_q[num]
        eq = en_q.get(num)
        tq = ta_q.get(num)

        notes = []
        if "[NEEDS REVIEW" in sq.get("note", "") or "[auto-" in sq.get("note", ""):
            notes.append(f"SI: {sq['note'].strip()}")
        if en_q and eq is None:
            notes.append("No matching question number found in English PDF")
        if ta_q and tq is None:
            notes.append("No matching question number found in Tamil PDF")

        row = {
            "AL/OL": level or "",
            "Subject": subject or "",
            "Sinhala Question": question_html(sq),
            "Sinhala Option 1": option_html(sq, "1"),
            "Sinhala Option 2": option_html(sq, "2"),
            "Sinhala Option 3": option_html(sq, "3"),
            "Sinhala Option 4": option_html(sq, "4"),
            "Sinhala Option 5": option_html(sq, "5"),
            "Sinhala Explanation": "",
            "English Question": question_html(eq) if eq else "",
            "English Option 1": option_html(eq, "1") if eq else "",
            "English Option 2": option_html(eq, "2") if eq else "",
            "English Option 3": option_html(eq, "3") if eq else "",
            "English Option 4": option_html(eq, "4") if eq else "",
            "English Option 5": option_html(eq, "5") if eq else "",
            "English Explanation": "",
            "Tamil Question": question_html(tq) if tq else "",
            "Tamil Option 1": option_html(tq, "1") if tq else "",
            "Tamil Option 2": option_html(tq, "2") if tq else "",
            "Tamil Option 3": option_html(tq, "3") if tq else "",
            "Tamil Option 4": option_html(tq, "4") if tq else "",
            "Tamil Option 5": option_html(tq, "5") if tq else "",
            "Tamil Explanation": "",
            "Type": "Other",
            "Answer": "",
            "Difficulty Level": "",
            "Sinhala Module": "",
            "Sinhala sub module": "",
            "Sinhala sub sub module": "",
            "Tags": "",
            "Notes": "; ".join(notes),
        }
        rows.append(row)
    return rows


def build_template_excel(rows):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    SINHALA_FONT = "Iskoola Pota"
    HEADER_FILL = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
    HEADER_FONT = Font(name=SINHALA_FONT, size=10, bold=True, color="FFFFFF")
    BODY_FONT = Font(name=SINHALA_FONT, size=10)
    THIN = Side(style="thin", color="D9D9D9")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    ws.append(TEMPLATE_HEADERS)
    for c in range(1, len(TEMPLATE_HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    for r, row in enumerate(rows, start=2):
        for c, header in enumerate(TEMPLATE_HEADERS, start=1):
            cell = ws.cell(row=r, column=c, value=row[header])
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = 60

    for c in range(1, len(TEMPLATE_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 20
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


app = Flask(__name__)

MAX_UPLOAD_MB = 15
app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_MB * 1024 * 1024

PAGE = """<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Sinhala MCQ PDF to Excel</title>
<style>
  :root { color-scheme: light; }
  * { box-sizing: border-box; }
  body {
    font-family: -apple-system, "Segoe UI", Roboto, Arial, sans-serif;
    max-width: 720px; margin: 40px auto; padding: 0 20px; color: #1a1a1a;
    background: #fafafa;
  }
  h1 { font-size: 22px; margin-bottom: 4px; }
  p.sub { color: #555; margin-top: 0; }
  .tabs { display: flex; gap: 8px; margin: 20px 0 16px; }
  .tab-btn {
    flex: 1; padding: 10px 14px; border: 1px solid #ccc; border-radius: 8px;
    background: #fff; cursor: pointer; font-size: 14px; font-weight: 600;
    color: #444;
  }
  .tab-btn.active { background: #1F4E78; color: #fff; border-color: #1F4E78; }
  .panel { display: none; }
  .panel.active { display: block; }
  .drop {
    border: 2px dashed #999; border-radius: 10px; padding: 28px; text-align: center;
    background: #fff; cursor: pointer; transition: border-color .15s, background .15s;
    font-size: 14px; color: #444;
  }
  .drop.drag { border-color: #1F4E78; background: #eef4fa; }
  .drop small { display: block; color: #888; margin-top: 6px; }
  .filename { margin-top: 8px; font-size: 13px; color: #1F4E78; font-weight: 600; word-break: break-all; }
  label.field { display: block; margin: 14px 0 4px; font-size: 13px; font-weight: 600; color: #333; }
  input[type=text] {
    width: 100%; padding: 8px 10px; border: 1px solid #ccc; border-radius: 6px; font-size: 14px;
  }
  button.go {
    margin-top: 18px; width: 100%; padding: 11px; border: none; border-radius: 8px;
    background: #1F4E78; color: #fff; font-size: 15px; font-weight: 600; cursor: pointer;
  }
  button.go:disabled { background: #99b3c4; cursor: not-allowed; }
  #status { margin-top: 14px; font-size: 13.5px; white-space: pre-wrap; line-height: 1.5; }
  #status.err { color: #b00020; }
  #status.ok { color: #1a7a33; }
  .hint { font-size: 12px; color: #888; margin-top: 22px; line-height: 1.5; }
</style>
</head>
<body>

<h1>Sinhala MCQ PDF &rarr; Excel</h1>
<p class="sub">Drop in an A/L-style MCQ PDF (FM Abhaya / FM Samantha / FM Ganganee fonts) and get a clean Unicode Excel file back.</p>

<div class="tabs">
  <button class="tab-btn active" data-tab="simple">Simple (one language)</button>
  <button class="tab-btn" data-tab="template">Template (Sinhala/English/Tamil)</button>
</div>

<div class="panel active" id="panel-simple">
  <div class="drop" id="drop-simple">
    Drop PDF here, or click to choose a file
    <small>Only .pdf is accepted &mdash; always more accurate than a .md/.txt export</small>
    <div class="filename" id="fn-simple"></div>
  </div>
  <input type="file" id="file-simple" accept="application/pdf" style="display:none">
  <button class="go" id="go-simple" disabled>Convert to Excel</button>
</div>

<div class="panel" id="panel-template">
  <label class="field">Sinhala PDF (required)</label>
  <div class="drop" id="drop-si">
    Drop Sinhala PDF here, or click to choose
    <div class="filename" id="fn-si"></div>
  </div>
  <input type="file" id="file-si" accept="application/pdf" style="display:none">

  <label class="field">English PDF (optional, same exam)</label>
  <div class="drop" id="drop-en">
    Drop English PDF here, or click to choose
    <div class="filename" id="fn-en"></div>
  </div>
  <input type="file" id="file-en" accept="application/pdf" style="display:none">

  <label class="field">Tamil PDF (optional, same exam)</label>
  <div class="drop" id="drop-ta">
    Drop Tamil PDF here, or click to choose
    <div class="filename" id="fn-ta"></div>
  </div>
  <input type="file" id="file-ta" accept="application/pdf" style="display:none">

  <label class="field">Level (e.g. AL / OL)</label>
  <input type="text" id="level" placeholder="AL">

  <label class="field">Subject</label>
  <input type="text" id="subject" placeholder="Biology">

  <button class="go" id="go-template" disabled>Convert to Excel</button>
</div>

<div id="status"></div>

<div class="hint">
  Nothing you upload is stored &mdash; each file is processed in memory for this
  one request and discarded. Module/sub-module, Tags, Difficulty Level and
  Answer are always left blank for you to fill in by hand (not something a
  script should guess). Anything the parser had to guess about is flagged in
  the Notes column &mdash; search for "NEEDS REVIEW" after each run.
</div>

<script>
function wireDrop(dropId, inputId, fnId, onChange) {
  const drop = document.getElementById(dropId);
  const input = document.getElementById(inputId);
  const fn = document.getElementById(fnId);
  drop.addEventListener('click', () => input.click());
  input.addEventListener('change', () => {
    if (input.files[0]) fn.textContent = input.files[0].name;
    onChange();
  });
  ['dragenter', 'dragover'].forEach(ev => drop.addEventListener(ev, e => {
    e.preventDefault(); drop.classList.add('drag');
  }));
  ['dragleave', 'drop'].forEach(ev => drop.addEventListener(ev, e => {
    e.preventDefault(); drop.classList.remove('drag');
  }));
  drop.addEventListener('drop', e => {
    if (e.dataTransfer.files[0]) {
      input.files = e.dataTransfer.files;
      fn.textContent = input.files[0].name;
      onChange();
    }
  });
}

document.querySelectorAll('.tab-btn').forEach(btn => {
  btn.addEventListener('click', () => {
    document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
    document.querySelectorAll('.panel').forEach(p => p.classList.remove('active'));
    btn.classList.add('active');
    document.getElementById('panel-' + btn.dataset.tab).classList.add('active');
    setStatus('');
  });
});

const goSimple = document.getElementById('go-simple');
wireDrop('drop-simple', 'file-simple', 'fn-simple', () => {
  goSimple.disabled = !document.getElementById('file-simple').files[0];
});

const goTemplate = document.getElementById('go-template');
wireDrop('drop-si', 'file-si', 'fn-si', updateTemplateBtn);
wireDrop('drop-en', 'file-en', 'fn-en', updateTemplateBtn);
wireDrop('drop-ta', 'file-ta', 'fn-ta', updateTemplateBtn);
function updateTemplateBtn() {
  goTemplate.disabled = !document.getElementById('file-si').files[0];
}

function setStatus(msg, cls) {
  const el = document.getElementById('status');
  el.textContent = msg;
  el.className = cls || '';
}

async function downloadResult(resp, fallbackName) {
  if (!resp.ok) {
    let msg = 'Conversion failed (HTTP ' + resp.status + ').';
    try { const j = await resp.json(); if (j.error) msg = j.error; } catch (e) {}
    throw new Error(msg);
  }
  const blob = await resp.blob();
  const disposition = resp.headers.get('Content-Disposition') || '';
  const match = disposition.match(/filename="?([^"]+)"?/);
  const name = match ? match[1] : fallbackName;
  const url = URL.createObjectURL(blob);
  const a = document.createElement('a');
  a.href = url; a.download = name;
  document.body.appendChild(a); a.click(); a.remove();
  URL.revokeObjectURL(url);
}

goSimple.addEventListener('click', async () => {
  const file = document.getElementById('file-simple').files[0];
  if (!file) return;
  goSimple.disabled = true;
  setStatus('Converting ' + file.name + ' ...');
  try {
    const fd = new FormData();
    fd.append('pdf', file);
    const resp = await fetch('/api/convert', { method: 'POST', body: fd });
    const flaggedHeader = resp.headers.get('X-Flagged-Questions');
    await downloadResult(resp, file.name.replace(/\\.pdf$/i, '') + '.xlsx');
    let msg = 'Done. Excel file downloaded.';
    if (flaggedHeader) msg += '\\nDouble-check these question numbers in the Notes column: ' + flaggedHeader;
    setStatus(msg, 'ok');
  } catch (e) {
    setStatus(e.message, 'err');
  } finally {
    goSimple.disabled = false;
  }
});

goTemplate.addEventListener('click', async () => {
  const si = document.getElementById('file-si').files[0];
  if (!si) return;
  goTemplate.disabled = true;
  setStatus('Converting ...');
  try {
    const fd = new FormData();
    fd.append('si', si);
    const en = document.getElementById('file-en').files[0];
    const ta = document.getElementById('file-ta').files[0];
    if (en) fd.append('en', en);
    if (ta) fd.append('ta', ta);
    fd.append('level', document.getElementById('level').value);
    fd.append('subject', document.getElementById('subject').value);
    const resp = await fetch('/api/convert-template', { method: 'POST', body: fd });
    const flaggedHeader = resp.headers.get('X-Flagged-Rows');
    await downloadResult(resp, si.name.replace(/\\.pdf$/i, '') + '.xlsx');
    let msg = 'Done. Excel file downloaded.';
    if (flaggedHeader) msg += '\\nSee the Notes column for these row numbers: ' + flaggedHeader;
    setStatus(msg, 'ok');
  } catch (e) {
    setStatus(e.message, 'err');
  } finally {
    goTemplate.disabled = false;
  }
});
</script>
</body>
</html>
"""


@app.route("/", methods=["GET"])
def home():
    return Response(PAGE, mimetype="text/html")


@app.route("/api/convert", methods=["POST"])
def convert_simple():
    f = request.files.get("pdf")
    if f is None or not f.filename:
        return {"error": "No PDF uploaded."}, 400
    if not f.filename.lower().endswith(".pdf"):
        return {"error": "Please upload a .pdf file."}, 400

    try:
        pdf_bytes = f.read()
        questions, order = extract_questions_from_pdf_bytes(pdf_bytes)
        if not order:
            return {"error": "No numbered questions (1., 2., 3. ...) were found in this PDF."}, 422
        xlsx_bytes = build_simple_excel(questions, order)
    except Exception as e:
        return {"error": f"Conversion failed: {e}"}, 500

    flagged = [n for n in order if "[NEEDS REVIEW" in questions[n]["note"] or "[auto-" in questions[n]["note"]]
    out_name = f.filename.rsplit(".", 1)[0] + ".xlsx"
    resp = send_file(
        io.BytesIO(xlsx_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=out_name,
    )
    if flagged:
        resp.headers["X-Flagged-Questions"] = str(flagged)
        resp.headers["Access-Control-Expose-Headers"] = "X-Flagged-Questions, Content-Disposition"
    return resp


@app.route("/api/convert-template", methods=["POST"])
def convert_template():
    si_f = request.files.get("si")
    en_f = request.files.get("en")
    ta_f = request.files.get("ta")
    level = request.form.get("level", "")
    subject = request.form.get("subject", "")

    if si_f is None or not si_f.filename:
        return {"error": "A Sinhala PDF is required."}, 400
    for f, label in ((si_f, "Sinhala"), (en_f, "English"), (ta_f, "Tamil")):
        if f is not None and f.filename and not f.filename.lower().endswith(".pdf"):
            return {"error": f"{label} file must be a .pdf."}, 400

    try:
        si_q, si_order = extract_questions_from_pdf_bytes(si_f.read())
        en_q, _ = extract_questions_from_pdf_bytes(en_f.read()) if (en_f and en_f.filename) else ({}, [])
        ta_q, _ = extract_questions_from_pdf_bytes(ta_f.read()) if (ta_f and ta_f.filename) else ({}, [])
        if not si_order:
            return {"error": "No numbered questions (1., 2., 3. ...) were found in the Sinhala PDF."}, 422
        rows = build_template_rows(si_q, si_order, en_q, ta_q, level, subject)
        xlsx_bytes = build_template_excel(rows)
    except Exception as e:
        return {"error": f"Conversion failed: {e}"}, 500

    flagged = [i + 1 for i, r in enumerate(rows) if r["Notes"]]
    out_name = si_f.filename.rsplit(".", 1)[0] + ".xlsx"
    resp = send_file(
        io.BytesIO(xlsx_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=out_name,
    )
    if flagged:
        resp.headers["X-Flagged-Rows"] = str(flagged)
        resp.headers["Access-Control-Expose-Headers"] = "X-Flagged-Rows, Content-Disposition"
    return resp


if __name__ == "__main__":
    app.run(debug=True, port=5000)
