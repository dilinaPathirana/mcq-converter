"""
Core Sinhala legacy-font MCQ PDF -> structured-question extraction logic.

PDF-only (the web app never accepts .md, so the heuristic fallback and its
english_words dependency are dropped here to keep the Vercel function bundle
small -- see the desktop tool's convert_mcq.py for the .md fallback path).
"""

import re

SINHALA_FONTS = {"FMAbhayax", "FMSamanthax", "FMGanganeex", "FMAbhaya", "FMBindumathi"}

SUP_START = "\x01"
SUP_END = "\x02"

DEGREE_FIX_RE = re.compile(r"(\d+)0([CF])(?![a-zA-Z0-9])")


def fix_degree_symbol(text):
    return DEGREE_FIX_RE.sub(r"\1" + chr(0xB0) + r"\2", text)


def split_superscript_runs(text):
    if SUP_START not in text:
        return [(False, text)] if text else []
    parts = []
    buf = ""
    in_sup = False
    for ch in text:
        if ch == SUP_START:
            if buf:
                parts.append((in_sup, buf))
            buf = ""
            in_sup = True
        elif ch == SUP_END:
            if buf:
                parts.append((in_sup, buf))
            buf = ""
            in_sup = False
        else:
            buf += ch
    if buf:
        parts.append((in_sup, buf))
    return parts


def strip_superscript_markers(text):
    return text.replace(SUP_START, "").replace(SUP_END, "")


def extract_from_pdf_bytes(pdf_bytes):
    import fitz
    from pandukabhaya import Converter

    conv = Converter("fm_abhaya")
    pua_re = re.compile("[" + chr(0xE000) + "-" + chr(0xF8FF) + "]")

    def convert_text(text, font):
        return conv.convert(text) if font in SINHALA_FONTS else text

    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    lines = []
    for page in doc:
        d = page.get_text("dict")
        blocks = sorted(d["blocks"], key=lambda b: (round(b["bbox"][1], 1), b["bbox"][0]))
        for b in blocks:
            for l in b.get("lines", []):
                spans = l.get("spans", [])
                n = len(spans)
                parts = []
                i = 0
                while i < n:
                    s = spans[i]
                    # See convert_mcq.py (desktop tool) for why this excludes
                    # Sinhala-font spans: the superscript flag is occasionally
                    # set on normal, full-size legacy-font text too (a table
                    # layout artifact), so only trust it for plain Latin spans.
                    is_sup = bool(s.get("flags", 0) & 1) and s["font"] not in SINHALA_FONTS
                    raw = s["text"]
                    text = convert_text(raw, s["font"])
                    if is_sup and raw.strip() == "0":
                        nxt = spans[i + 1]["text"] if i + 1 < n else ""
                        if nxt.lstrip().startswith(("C", "F")):
                            parts.append(chr(0xB0))
                            i += 1
                            continue
                    if is_sup and text.strip():
                        parts.append(SUP_START + text + SUP_END)
                    else:
                        parts.append(text)
                    i += 1
                line_text = pua_re.sub("", "".join(parts))
                if line_text.strip():
                    lines.append(line_text)
            lines.append("")
    doc.close()
    return fix_degree_symbol("\n".join(lines))


NEW_LINE_RE = re.compile(r"^(\d{1,2}\.|\(\d\)|\([A-E]\)|[A-E]\s*$|[A-E]\s*-)")


def reflow(raw_text):
    lines = raw_text.split("\n")
    logical, buf = [], ""
    for l in lines:
        s = l.strip()
        if s == "":
            if buf:
                logical.append(buf)
                buf = ""
            continue
        if NEW_LINE_RE.match(s):
            if buf:
                logical.append(buf)
            buf = s
        else:
            buf = f"{buf} {s}" if buf else s
    if buf:
        logical.append(buf)
    fixed = []
    for l in logical:
        if len(re.findall(r"\(\d\)", l)) >= 2:
            fixed.extend(p.strip() for p in re.split(r"(?=\(\d\)\s)", l) if p.strip())
        else:
            fixed.append(l)
    return fixed


Q_START_RE = re.compile(r"^(\d{1,2})\.\s*(.*)$")
OPT_NUM_RE = re.compile(r"^\((\d)\)\s*(.*)$")
OPT_LET_RE = re.compile(r"^\(([A-E])\)\s*(.*)$")


def new_q(num):
    return {"num": num, "stem_lines": [], "num_opts": {}, "let_opts": {}, "extra": [], "note": ""}


def find_combo_note(raw_text):
    m = re.search(r"\d{1,2}\s*[-" + chr(0x2013) + r"]\s*\d{1,2}[^\n]{0,60}?(?:" + chr(0x0db4) + chr(0x0daf) + chr(0x0db1) + chr(0x0db8) + "|based)", raw_text)
    if not m:
        return raw_text, None
    nums = re.match(r"(\d{1,2})\s*[-" + chr(0x2013) + r"]\s*(\d{1,2})", raw_text[m.start():])
    lo, hi = int(nums.group(1)), int(nums.group(2))
    tail = raw_text[m.start():m.start() + 400]
    m2 = re.search(r"\n\s*\d{1,2}\.\s", tail[10:])
    if m2:
        tail = tail[:10 + m2.start()]
    note = re.sub(r"\s+", " ", tail).strip()
    cleaned = raw_text[:m.start()] + raw_text[m.start() + len(tail):]
    return cleaned, (lo, hi, note)


def parse_questions(lines, max_q=60, combo_note=None):
    start = 0
    for i, l in enumerate(lines):
        m = Q_START_RE.match(l)
        if m and int(m.group(1)) == 1:
            start = i
            break
    lines = lines[start:]

    questions = {}
    order = []
    cur = None

    i = 0
    while i < len(lines):
        l = lines[i]
        m = Q_START_RE.match(l)
        if m and 1 <= int(m.group(1)) <= max_q:
            num = int(m.group(1))
            cur = new_q(num)
            questions[num] = cur
            order.append(num)
            rest = m.group(2).strip()
            if rest:
                cur["stem_lines"].append(rest)
            i += 1
            continue
        mo = OPT_NUM_RE.match(l)
        if mo and cur:
            key = mo.group(1)
            if key in cur["num_opts"] and cur["num_opts"][key] != mo.group(2).strip():
                prev_idx = order.index(cur["num"]) - 1
                prev_num = order[prev_idx] if prev_idx >= 0 else None
                if prev_num is not None and not questions[prev_num]["num_opts"]:
                    questions[prev_num]["num_opts"] = dict(cur["num_opts"])
                    questions[prev_num]["note"] += " [auto-recovered option table from layout - please verify]"
                cur["num_opts"] = {}
            cur["num_opts"][key] = mo.group(2).strip()
            i += 1
            continue
        mol = OPT_LET_RE.match(l)
        if mol and cur:
            cur["let_opts"][mol.group(1)] = mol.group(2).strip()
            i += 1
            continue
        if cur:
            cur["extra"].append(l)
        i += 1

    for num, q in questions.items():
        extra_q = list(q["extra"])

        letter_lines = [e for e in extra_q if re.match(r"^[A-E]\s*-\s*", e)]
        if letter_lines:
            rest = [e for e in extra_q if e not in letter_lines]
            q["premises"] = letter_lines
            if rest:
                q["stem_lines"] = [rest[-1]]
                extra_q = rest[:-1]
            else:
                extra_q = []
            q["note"] += " [auto-detected premise list - please verify]"

        if extra_q and q["num_opts"]:
            keys = sorted(q["num_opts"].keys(), key=int)
            for k in keys:
                val = q["num_opts"][k]
                guard = 0
                while val and not strip_superscript_markers(val).rstrip().endswith((".", "?", "!")) and extra_q and guard < 5:
                    val = val.rstrip() + " " + extra_q.pop(0)
                    guard += 1
                q["num_opts"][k] = val
            if extra_q:
                q["stem_lines"] += extra_q
            extra_q = []
        q["extra"] = extra_q

        if q["let_opts"] and not q["num_opts"]:
            q["premises"] = [f"({k}) {v}" for k, v in sorted(q["let_opts"].items())]
            if combo_note and combo_note[0] <= num <= combo_note[1]:
                q["note"] += " [combo-answer question - see Notes for shared answer key]"
                q["notes_full"] = combo_note[2]
            else:
                q["note"] += " [NEEDS REVIEW: lettered options with no numbered answer choices found]"

        if not q["num_opts"] and not q["let_opts"]:
            q["note"] += " [NEEDS REVIEW: no options detected]"

    return questions, order


def extract_questions_from_pdf_bytes(pdf_bytes, max_q=60):
    """Full pipeline: PDF bytes -> (questions dict, order list)."""
    raw = extract_from_pdf_bytes(pdf_bytes)
    raw, combo_note = find_combo_note(raw)
    lines = reflow(raw)
    return parse_questions(lines, max_q=max_q, combo_note=combo_note)


def _rich_or_plain(text):
    if SUP_START not in text:
        return text
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont

    sup_font = InlineFont(vertAlign="superscript")
    runs = []
    for is_sup, chunk in split_superscript_runs(text):
        if is_sup:
            runs.append(TextBlock(sup_font, chunk))
        else:
            runs.append(chunk)
    return CellRichText(runs)


def build_simple_excel(questions, order):
    """Build the plain Q No / Question / Option 1-5 / Notes workbook.
    Returns raw xlsx bytes."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "MCQ"

    SINHALA_FONT = "Iskoola Pota"
    HEADER_FILL = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
    HEADER_FONT = Font(name=SINHALA_FONT, size=11, bold=True, color="FFFFFF")
    BODY_FONT = Font(name=SINHALA_FONT, size=11)
    THIN = Side(style="thin", color="D9D9D9")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    headers = ["Q No", "Question", "Option 1", "Option 2", "Option 3", "Option 4", "Option 5", "Notes"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for i, w in enumerate([6, 55, 22, 22, 22, 22, 22, 45], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for num in order:
        q = questions[num]
        stem = " ".join(q["stem_lines"]).strip()
        question_text = stem
        if q.get("premises"):
            question_text += "\n" + "\n".join(q["premises"])
        opts = q["num_opts"]
        note = q.get("notes_full", "") or q.get("note", "").strip()
        row_values = [num, question_text, opts.get("1", ""), opts.get("2", ""), opts.get("3", ""),
                      opts.get("4", ""), opts.get("5", ""), note]
        row_values = [_rich_or_plain(v) if isinstance(v, str) else v for v in row_values]
        ws.append(row_values)
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(
                horizontal="center" if c == 1 else "left", vertical="top", wrap_text=True
            )
        ws.row_dimensions[r].height = 60

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Multi-language HTML template mode (mirrors convert_mcq_template.py)
# ---------------------------------------------------------------------------

TEMPLATE_HEADERS = [
    "AL/OL", "Subject",
    "Sinhala Question", "Sinhala Option 1", "Sinhala Option 2", "Sinhala Option 3",
    "Sinhala Option 4", "Sinhala Option 5", "Sinhala Explanation",
    "English Question", "English Option 1", "English Option 2", "English Option 3",
    "English Option 4", "English Option 5", "English Explanation",
    "Tamil Question", "Tamil Option 1", "Tamil Option 2", "Tamil Option 3",
    "Tamil Option 4", "Tamil Option 5", "Tamil Explanation",
    "Type", "Answer", "Difficulty Level",
    "Sinhala Module", "Sinhala sub module", "Sinhala sub sub module", "Tags",
    "Notes",
]


def sup_to_html(text):
    if SUP_START not in text:
        return text
    out = []
    for is_sup, chunk in split_superscript_runs(text):
        out.append(f"<sup>{chunk}</sup>" if is_sup else chunk)
    return "".join(out)


def html_wrap(text):
    text = sup_to_html((text or "").strip())
    return f"<p>{text}</p>" if text else ""


def question_html(q):
    parts = [html_wrap(s) for s in q.get("stem_lines", []) if s.strip()]
    premises = q.get("premises")
    if premises:
        cleaned = []
        for p in premises:
            m = re.match(r"^\(?([A-E])\)?\s*-?\s*(.*)$", p)
            if m:
                cleaned.append(f"{m.group(1)} - {m.group(2)}")
            else:
                cleaned.append(p)
        parts.append(html_wrap(", ".join(cleaned)))
    return "".join(parts)


def option_html(q, key):
    return html_wrap(q.get("num_opts", {}).get(key, ""))


def build_template_rows(si_q, si_order, en_q, ta_q, level, subject):
    rows = []
    for num in si_order:
        sq = si_q[num]
        eq = en_q.get(num)
        tq = ta_q.get(num)

        notes = []
        if "[NEEDS REVIEW" in sq.get("note", "") or "[auto-" in sq.get("note", ""):
            notes.append(f"SI: {sq['note'].strip()}")
        if en_q and eq is None:
            notes.append("No matching question number found in English PDF")
        if ta_q and tq is None:
            notes.append("No matching question number found in Tamil PDF")

        row = {
            "AL/OL": level or "",
            "Subject": subject or "",
            "Sinhala Question": question_html(sq),
            "Sinhala Option 1": option_html(sq, "1"),
            "Sinhala Option 2": option_html(sq, "2"),
            "Sinhala Option 3": option_html(sq, "3"),
            "Sinhala Option 4": option_html(sq, "4"),
            "Sinhala Option 5": option_html(sq, "5"),
            "Sinhala Explanation": "",
            "English Question": question_html(eq) if eq else "",
            "English Option 1": option_html(eq, "1") if eq else "",
            "English Option 2": option_html(eq, "2") if eq else "",
            "English Option 3": option_html(eq, "3") if eq else "",
            "English Option 4": option_html(eq, "4") if eq else "",
            "English Option 5": option_html(eq, "5") if eq else "",
            "English Explanation": "",
            "Tamil Question": question_html(tq) if tq else "",
            "Tamil Option 1": option_html(tq, "1") if tq else "",
            "Tamil Option 2": option_html(tq, "2") if tq else "",
            "Tamil Option 3": option_html(tq, "3") if tq else "",
            "Tamil Option 4": option_html(tq, "4") if tq else "",
            "Tamil Option 5": option_html(tq, "5") if tq else "",
            "Tamil Explanation": "",
            "Type": "Other",
            "Answer": "",
            "Difficulty Level": "",
            "Sinhala Module": "",
            "Sinhala sub module": "",
            "Sinhala sub sub module": "",
            "Tags": "",
            "Notes": "; ".join(notes),
        }
        rows.append(row)
    return rows


def build_template_excel(rows):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    SINHALA_FONT = "Iskoola Pota"
    HEADER_FILL = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
    HEADER_FONT = Font(name=SINHALA_FONT, size=10, bold=True, color="FFFFFF")
    BODY_FONT = Font(name=SINHALA_FONT, size=10)
    THIN = Side(style="thin", color="D9D9D9")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    ws.append(TEMPLATE_HEADERS)
    for c in range(1, len(TEMPLATE_HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    for r, row in enumerate(rows, start=2):
        for c, header in enumerate(TEMPLATE_HEADERS, start=1):
            cell = ws.cell(row=r, column=c, value=row[header])
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = 60

    for c in range(1, len(TEMPLATE_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 20
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
